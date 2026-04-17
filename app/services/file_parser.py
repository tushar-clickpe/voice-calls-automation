"""File parser service - reads contacts from uploaded CSV/XLSX files."""

import csv
import io
import re
from typing import Optional
from openpyxl import load_workbook


def parse_uploaded_file(
    file_content: bytes,
    filename: str,
    phone_column: str = None,
    name_column: str = None,
) -> list[dict]:
    """Parse an uploaded CSV or XLSX file and extract contacts.

    Args:
        file_content: Raw bytes of the uploaded file
        filename: Original filename (used to detect format)
        phone_column: Explicit phone column name (auto-detected if None)
        name_column: Explicit name column name (auto-detected if None)

    Returns:
        List of dicts with 'phone', 'name', and any extra columns
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        records, headers = _parse_csv(file_content)
    elif ext in ("xlsx", "xls"):
        records, headers = _parse_xlsx(file_content)
    else:
        raise ValueError(
            f"Unsupported file format: .{ext}. Please upload a .csv or .xlsx file."
        )

    if not records:
        return []

    # Auto-detect columns
    phone_col = _find_column(
        headers, phone_column,
        ["phone", "mobile", "number", "contact", "tel", "cell", "whatsapp", "phone number", "mobile number"],
    )
    name_col = _find_column(
        headers, name_column,
        ["name", "customer", "client", "person", "full name", "customer name", "borrower"],
    )

    if not phone_col:
        raise ValueError(
            f"Could not auto-detect phone column. Available columns: {headers}. "
            "Please specify the phone column name."
        )

    contacts = []
    for row in records:
        phone = str(row.get(phone_col, "")).strip()
        if not phone:
            continue

        # Normalize phone
        phone = re.sub(r"[\s\-\(\)]", "", phone)
        # Remove .0 from numbers cast to float
        if phone.endswith(".0"):
            phone = phone[:-2]
        # Skip if it's not a plausible phone number (at least 7 digits)
        digits_only = re.sub(r"[^\d]", "", phone)
        if len(digits_only) < 7:
            continue

        contact = {
            "phone": phone,
            "name": str(row.get(name_col, "")).strip() if name_col else "",
        }

        # Include all other columns as extra data
        for key, value in row.items():
            if key not in (phone_col, name_col):
                contact[key] = str(value).strip() if value is not None else ""

        contacts.append(contact)

    return contacts


def get_file_headers(file_content: bytes, filename: str) -> list[str]:
    """Get just the column headers from a file."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        _, headers = _parse_csv(file_content, max_rows=0)
    elif ext in ("xlsx", "xls"):
        _, headers = _parse_xlsx(file_content, max_rows=0)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")
    return headers


def get_file_preview(file_content: bytes, filename: str, rows: int = 5) -> tuple[list[str], list[dict]]:
    """Get headers and first N rows from a file for preview."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        records, headers = _parse_csv(file_content, max_rows=rows)
    elif ext in ("xlsx", "xls"):
        records, headers = _parse_xlsx(file_content, max_rows=rows)
    else:
        raise ValueError(f"Unsupported file format: .{ext}")
    return headers, records


def _parse_csv(file_content: bytes, max_rows: int = None) -> tuple[list[dict], list[str]]:
    """Parse CSV file content into list of dicts."""
    # Try to decode with different encodings
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = file_content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode the CSV file. Please save it as UTF-8.")

    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    # Clean up headers
    headers = [h.strip() for h in headers if h and h.strip()]

    records = []
    for i, row in enumerate(reader):
        if max_rows is not None and i >= max_rows:
            break
        # Clean keys
        cleaned = {k.strip(): v for k, v in row.items() if k and k.strip()}
        records.append(cleaned)

    return records, headers


def _parse_xlsx(file_content: bytes, max_rows: int = None) -> tuple[list[dict], list[str]]:
    """Parse XLSX file content into list of dicts."""
    wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    # Filter out empty headers
    headers = [h for h in headers if h]

    records = []
    for i, row in enumerate(rows_iter):
        if max_rows is not None and i >= max_rows:
            break
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        record = {}
        for j, header in enumerate(headers):
            value = row[j] if j < len(row) else None
            record[header] = value if value is not None else ""
        records.append(record)

    wb.close()
    return records, headers


def _find_column(headers: list[str], explicit: str = None, keywords: list[str] = None) -> Optional[str]:
    """Find a column by explicit name or keyword matching."""
    if explicit:
        for h in headers:
            if h.lower().strip() == explicit.lower().strip():
                return h
        return None

    if keywords:
        for h in headers:
            h_lower = h.lower().strip()
            for kw in keywords:
                if kw in h_lower:
                    return h
    return None
